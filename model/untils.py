import openpyxl
import datetime



def read_data(file_name):
    """读取excel文件内容"""
    wb=openpyxl.load_workbook(file_name)
    ws=wb.active
    row=ws.max_row
    col=ws.max_column
    dates=[]
    for i in range(2,row+1):
        date=[]
        for j in range(1,col+1):
            date.append(str(ws.cell(i,j).value))
        dates.append(date)
    wb.close()
    return dates

def write_data(file_name,data,header):
    """写入excel"""
    wb=openpyxl.Workbook()
    ws=wb.active
    row=len(data)+1
    col=len(data[0])
    for index in range(len(header)):
        ws.cell(row=1,column=index+1,value=header[index])

    for i in range(2,row+1):
        for j in range(1,col+1):
            ws.cell(row=i,column=j,value=data[i-2][j-1])

    wb.save(file_name)
    wb.close()


if __name__ == '__main__':
    now=datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    data=[('a1','success','wu'),('a1','success','wu')]
    header=('钢卷号','jieguo','yuanxin')
    write_data(r'F:\autotest\peihuo\result_file\{}.xlsx'.format(now),3,3,data,header)